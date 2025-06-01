from datetime import datetime, timedelta
from typing import List, Optional
from collections import defaultdict


import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

from matplotlib import ticker


from database import DataCoin, User
from .name_transformer import transformer


def file_opener(file_name):
    # Открываем файл Excel с помощью openpyxl
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    row_count = ws.max_row-1
    total = 0

    # Проходимся по строкам и суммируем значения в столбце G
    for row in ws.iter_rows(min_row=2, max_col=12):
        if not row[9].value:
            continue

        if row[11].value != "Метка 13":
            # проверка на число
            if isinstance(row[9].value, (int, float)):
                total += row[9].value
            else:
                pass

    total_r = round(total, 2)
    return total_r, row_count



def more_info(file_name):
    # df = pd.read_excel(file_name)
    # countryroad = df[df.columns[0]].unique()  # эта переменная считает кол-во стран
    df = 0
    countryroad = 0
    sold = 0
    try:
        df = pd.read_excel(file_name)
        countryroad = df[df.columns[0]].unique()  # эта переменная считает кол-во стран
        # Получить сумму элементов в 7 столбце
        sold = df.iloc[:, 16].sum() #ФАЙЛ|||| Сумма продажи

    except Exception:
        print(datetime.now(), "| ", f"Ошибка открытия файла")

    return len(df), len(countryroad), sold


def countries(file_name):
    df = pd.read_excel(file_name)
    result = []
    grouped = df.groupby("Страна").size()
    for country, count in grouped.items():
        # result += f"{mydict1[country]} {str(count):<5}{country}\n            /{mydict[country]}\n"
        result.append(
            [
                transformer.get_country_code(country),  # Флаг страны
                count,  # Кол-во монет
                country,  # Русское название страны
                transformer.get_country_eng_short_name(
                    country
                ),  # Короткое англ. название
            ]
        )
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    count_euro = 0
    for row in ws.iter_rows(min_row=1, max_col=9):
        if "евро" in row[3].value:   #ФАЙЛ||||
            count_euro += 1

    result.append(
        [
            f"🇪🇺",
            count_euro,
            f"Евросоюз",
            f"Europe",
        ]
    )
    return result


def euro(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    euros = []

    for row in ws.iter_rows(min_row=1, max_col=20):  # ФАЙЛ |
        if "евро" in row[3].value:
            des2 = f"{row[4].value}г." if row[4].value else ""  # год
            des3 = (
                f"\nРазновидность: {transformer.get_coin_difference(row[5].value)}"
                if row[5].value
                else ""
            )  # монетный двор
            des4 = f"\n{row[6].value}" if row[6].value else ""  # Наименование
            des5 = f"\nМоя цена: {row[16].value} ₽" if row[16].value else ""  # Моя цена
            des6 = f"\nКомментарий: {row[19].value}" if row[19].value else ""  # Комментарий
            cena = f" {row[9].value} ₽" if row[9].value else ""  # Цена

            euros.append(
                [
                    row[0].value,
                    f"🇪🇺 {transformer.get_country_code(row[0].value)}",  # Страна
                    row[3].value,  # номинал
                    des2,  # ГОД
                    cena,
                    des3,  # монетный двор
                    des4,  # Наименование
                    des5,  # покупка
                    des6,  # комментарий
                ]
            )

    return euros


def strana(file_name: str, text_in: str):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    country_key = transformer.get_country_rus_name(text_in[1:])
    results = []

    argentina_varieties = {
        'A': 'ОМД: Тэджон, Южная Корея. 6-угольная звезда над датой',
        'B': 'ОМД: 6-гранный цветок над датой. Правильная надпись "PROVINCIAS"',
        'B-er': 'ОМД: 6-гранный цветок над датой. Ошибочная надпись "PROVINGIAS"',
        'C': 'ОМД: Париж, Франция. Ромашка с 6 лепестками над датой'
    }

    historical_countries = {
        "Российская империя", "Япония", "Швеция", "Османская империя",
        "Нидерланды", "Испания", "Греция"
    }

    for row in ws.iter_rows(min_row=1, max_col=19):
        if row[0].value != country_key:
            continue

        country = row[0].value
        year = f"{row[4].value}г." if row[4].value else ""
        mint = row[5].value
        name = row[6].value or ""
        my_price = f"Моя цена: {row[16].value} ₽" if row[16].value else ""
        comment = f"Комментарий: {row[18].value}" if row[18].value else ""
        price = f"{row[9].value} ₽" if row[9].value else ""

        if country == "Аргентина":
            variety = argentina_varieties.get(mint, "")
            desc3 = f"Разновидность: {variety}" if variety else ""
        elif country in historical_countries:
            desc3 = ""
            if mint:
                desc3 += f"Разновидность: {transformer.get_coin_difference(mint)}\n"
            if row[1].value:
                desc3 += f"Период: {row[1].value}"
        else:
            desc3 = f"Разновидность: {transformer.get_coin_difference(mint)}" if mint else ""

        results.append([
            country,
            transformer.get_country_code(country),
            row[3].value,  # номинал
            year,
            price,
            desc3,
            name,
            my_price,
            comment,
        ])

    return results


def func_swap(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    country_data = defaultdict(list)

    for row in ws.iter_rows(min_row=2, max_col=18):
        des2 = f"{row[2].value}г." if row[2].value else ""  # год
        desc4 = f"\n{row[4].value}" if row[4].value else ""  # Наименование
        desc3 = (
            f"\nРазновидность: {transformer.get_coin_difference(row[3].value)}"
            if row[3].value
            else ""
        )  # монетный двор
        desc10 = f"\nКомментарий: {row[10].value}" if row[10].value else ""  # комментарий

        country_data[row[0].value].append(
            [
                transformer.get_country_code(row[0].value),
                row[1].value,  # Номинал
                des2,  # ГОД
                f" {row[6].value} ₽",  # Цена
                f"\nКол-во: {row[7].value}",  # Кол-во
                desc3,  # монетный двор
                desc4,  # Наименование
                desc10,  # комментарий
            ]
        )

    return country_data


def get_top_10_coin(file_name, mode):
    df = pd.read_excel(file_name)
    arr = []
    df.fillna(value="", inplace=True)

    if mode == "old":
        df = df.sort_values(by="Год", ascending=True)
    elif mode == "novelty":
        df = df.sort_values(by="Год", ascending=False)
    elif mode == "expensive_value":
        df["Цена, RUB [uCoin]"] = pd.to_numeric(df["Цена, RUB [uCoin]"], errors="coerce")
        df = df.sort_values(by="Цена, RUB [uCoin]", ascending=False)
    elif mode == "cheap_value":
        df["Цена, RUB [uCoin]"] = pd.to_numeric(df["Цена, RUB [uCoin]"], errors="coerce")
        df = df.sort_values(by="Цена, RUB [uCoin]", ascending=True)
    elif mode == "last_append":
        df = df.sort_values(by="Добавлено", ascending=False)
    elif mode == "first_append":
        df = df.sort_values(by="Добавлено", ascending=True)

    top_10 = df.head(10)

    # Проходимся по строкам и суммируем значения в столбце G
    for row in top_10.iterrows():
        desc4 = f"\n{row[1][6]}" if row[1][6] else ""  # Наименование
        desc3 = f"{row[1][5]}" if row[1][5] else ""  # монетный двор

        original_date_str = str(row[1][13])
        original_date = datetime.strptime(original_date_str, '%Y-%m-%d %H:%M:%S')

        # Получаем название месяца на русском языке
        month_names = [
            'января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
            'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря'
        ]
        formatted_date = f"{original_date.day} {month_names[original_date.month - 1]} {original_date.year}"

        desc11 = f"\nДобавлено: {formatted_date}" if row[1][13] else ""  # дата добавления
        desc10 = f"\nКомментарий: {row[1][18]}" if row[1][18] else ""  # комментарий
        desc5 = f" {row[1][9]} ₽" if row[1][9] else ""  # Цена
        arr.append(
            [
                transformer.get_country_code(row[1][0]),  # Флаг
                f"{row[1][0]}\n",
                row[1][3],  # Номинал
                row[1][4],  # Год
                desc5,  # Цена
                desc3,  # монетный двор
                desc4,  # Наименование
                desc10,  # комментарий
                desc11, #Дата добавления
            ]
        )
    return arr


def get_fig_width(data_length: int) -> int:
    if data_length > 90:
        return 20
    return 15 * (data_length // 60 or 1)


def get_date_annotation(date_value: str, data_length: int) -> str:
    date_ = datetime.strptime(date_value, "%Y.%m.%d")

    if data_length > 350:
        date_str = date_.strftime("%b %Y")
    elif data_length > 30 * 6:
        date_str = date_.strftime("%d %b")
    else:
        date_str = date_.strftime("%d.%m")

    return date_str


def get_fig_marker(data_length: int) -> str:
    if data_length < 61:
        return "o"
    return ""


def get_graph(tg_id, limit: Optional[int] = 30):
    owne1r = User.get(tg_id)

    len_active = len(DataCoin.get_for_user(tg_id, limit))

    graph_coin_data: List[DataCoin] = DataCoin.get_for_user(tg_id, limit)
    graph_date = []
    graph_sum = []
    graph_coin_count = []

    last_date = datetime.now().date()

    for sublist in graph_coin_data[::1]:
        while datetime.strptime(sublist.datetime, "%Y.%m.%d").date() != last_date:
            graph_date.append(last_date.strftime("%Y.%m.%d"))
            graph_sum.append(None)
            graph_coin_count.append(None)
            last_date -= timedelta(days=1)

        graph_date.append(sublist.datetime)
        graph_sum.append(sublist.totla_sum)
        graph_coin_count.append(sublist.totla_count)
        last_date -= timedelta(days=1)

    if limit:
        graph_date = graph_date[:limit]
        graph_sum = graph_sum[:limit]
        graph_coin_count = graph_coin_count[:limit]

    data_length = len(graph_date)

    step = data_length // 15 or 1

    fig_height = 10
    fig_width = get_fig_width(data_length)
    fig_dpi = 100

    plt.clf()
    fig, ax1 = plt.subplots(figsize=(fig_width, fig_height), dpi=fig_dpi)

    ax2 = ax1.twinx()

    ax1.plot(
        graph_date[::-1],
        graph_sum[::-1],
        marker=get_fig_marker(data_length),
        color='#0698FE',
        markersize=5,
    )
    ax2.plot(
        graph_date[::-1],
        graph_coin_count[::-1],
        marker=get_fig_marker(data_length),
        markersize=3,
        color='#30BA8F',
        linewidth=0.7
    )

    filtered_sum = [x for x in graph_sum if x is not None]  # filtered_sum равно [10, 20, 40, 50]
    maxim = max(filtered_sum)
    mimin = min(filtered_sum)
    average = sum(filtered_sum) / len(filtered_sum)
    average = round(average, 2)
    last = graph_sum[0]
    date = datetime.now
    date1 = date().strftime("%d.%m.%Y %H:%M")

    filtered_count = [x for x in graph_coin_count if x is not None]
    max_count = max(filtered_count)
    min_count = min(filtered_count)
    raznica = max_count-min_count
    y_min = min_count - 2
    y_max = min_count + raznica+2
    ax2.set_ylim(y_min, y_max)

    date_without_year = list(map(lambda value: get_date_annotation(value, data_length), graph_date))

    ax2.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    plt.xticks(graph_date[::step], date_without_year[::step])

    plt.title("Стоимость коллекции, руб")

    # Добавить текст на график
    plt.text(0, 1.07, " {}".format(owne1r.user_name), transform=plt.gca().transAxes)
    plt.text(0, 1.05, " {}".format(date1), transform=plt.gca().transAxes)

    plt.text(-0.08, 1.02, "Стоимость", color="#0698FE",  transform=plt.gca().transAxes)
    plt.text(1.02, 1.04, "Кол-во", color="#30BA8F", transform=plt.gca().transAxes)
    plt.text(1.02, 1.02, "монет", color="#30BA8F", transform=plt.gca().transAxes)

    plt.text(
        0,
        -0.1,
        "[◉_◉] Минимум = {} р.".format(mimin),
        color="red",
        transform=plt.gca().transAxes
    )
    plt.text(
        0.2,
        -0.1,
        "(◕‿◕) Максимум = {} р.".format(maxim),
        color="green",
        transform=plt.gca().transAxes,
    )
    plt.text(
        0.4,
        -0.1,
        "(─‿‿─) Средняя = {} р.".format(average),
        color="brown",
        transform=plt.gca().transAxes,
    )
    plt.text(
        0.8,
        -0.1,
        "(• ◡•) Последняя = {} р.".format(last),
        color="blue",
        transform=plt.gca().transAxes,
    )

    #  Прежде чем рисовать вспомогательные линии
    #  необходимо включить второстепенные деления
    ax1.minorticks_on()

    #  Определяем внешний вид линий основной сетки:
    ax1.grid(which="major")

    #  Определяем внешний вид линий вспомогательной
    #  сетки:
    ax1.grid(
        which="minor",
        linestyle=":",
    )

    path = f"./users_files/{owne1r.user_coin_id}_grafik.png"
    plt.savefig(path)

    return path, len_active


